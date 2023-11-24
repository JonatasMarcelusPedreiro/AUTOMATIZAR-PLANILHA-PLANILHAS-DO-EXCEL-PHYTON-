#Para come�ar, precisamos instalar a biblioteca openpyxl, que � a principal biblioteca para manipula��o de arquivos Excel em Python.

#Para instalar, basta executar o seguinte comando:


#pip install openpyxl
#Com a biblioteca instalada, podemos come�ar a escrever c�digo para automatizar tarefas no Excel.


#Vamos come�ar criando uma nova planilha em branco com o nome �Exemplo.xlsx�.

#Para fazer isso, utilizaremos a seguinte sequ�ncia de comandos em Python:


from openpyxl import Workbook

# Cria uma nova planilha em branco
wb = Workbook()

# Seleciona a planilha ativa
planilha = wb.active

# Define o valor da c�lula A1
planilha['A1'] = 'Ol�, Excel com Python!'

# Salva o arquivo
wb.save('Exemplo.xlsx')


#Neste exemplo, importamos a classe Workbook da biblioteca openpyxl.
#Em seguida, criamos um objeto wb do tipo Workbook, que representa a planilha em branco.
#Em seguida, selecionamos a planilha ativa usando o atributo active.
#Depois disso, definimos o valor da c�lula A1 da planilha usando a nota��o de �ndice 'A1'.
#No exemplo, atribu�mos o texto Ol�, Excel com Python! para a c�lula A1.
#Por fim, salvamos o arquivo com o nome �Exemplo.xlsx� usando o m�todo save do objeto wb.


#Al�m de criar uma nova planilha, tamb�m podemos usar o Python para ler dados de uma planilha existente.

#Para fazer isso, utilizamos a seguinte sequ�ncia de comandos:

from openpyxl import load_workbook

# Carrega o arquivo existente
wb = load_workbook('Exemplo.xlsx')

# Seleciona a planilha ativa
planilha = wb.active

# L� o valor da c�lula A1
valor_celula = planilha['A1'].value

# Imprime o valor na tela
print(valor_celula)


#Neste exemplo, importamos a fun��o load_workbook da biblioteca openpyxl.
#Em seguida, carregamos o arquivo �Exemplo.xlsx� usando a fun��o load_workbook e atribu�mos o objeto resultante � vari�vel wb.
#Depois disso, selecionamos a planilha ativa utilizando o atributo active.
#Em seguida, lemos o valor da c�lula A1, utilizando a nota��o de �ndice 'A1', e atribu�mos o valor � vari�vel valor_celula.
#Por fim, imprimimos o valor da c�lula A1 na tela utilizando a fun��o print.






#Atualizando dados em uma planilha existente
#Al�m de ler dados de uma planilha existente, tamb�m podemos atualizar os dados usando o Python.

#Para isso, utilizamos a seguinte sequ�ncia de comandos:


from openpyxl import load_workbook

# Carrega o arquivo existente
wb = load_workbook('Exemplo.xlsx')

# Seleciona a planilha ativa
planilha = wb.active

# Atualiza o valor da c�lula A1
planilha['A1'] = 'Novo valor'

# Salva o arquivo
wb.save('Exemplo.xlsx')

#Neste exemplo, importamos a fun��o load_workbook da biblioteca openpyxl.
#Em seguida, carregamos o arquivo �Exemplo.xlsx� usando a fun��o load_workbook e atribu�mos o objeto resultante � vari�vel wb.
#Depois disso, selecionamos a planilha ativa utilizando o atributo active.
#Em seguida, atualizamos o valor da c�lula A1, utilizando a nota��o de �ndice 'A1' e atribu�mos o novo valor.
#Por fim, salvamos o arquivo usando o m�todo save do objeto wb.
